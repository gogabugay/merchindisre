import re
from openpyxl import load_workbook
from os import listdir
photos=listdir('/home/yuriy/Документы/python/Виктория')     #бахаем список файлов
photos2=[]
for f in photos:
	f=f.lower()
	exclude=['виктория','виктория. ', '.jpg', '.jpeg','шоссе', 'ул.','б-р', 'д  ', 'бвар']
	for n in exclude:
		f=re.sub(n,'',f)
	f=f.lstrip()
	f=re.sub(('(\d)'), '',f)
	'''if f[:11]=='профсоюзная':
		f=f[:24]
	else:
		f=f[:13]'''
	if f=='ак.анохина 2.':
		f='академика ано'
	elif f=='б-р дм донско':
		f='дмитрия донск'
	elif f=='каширское шос':
		f='домодедово г,'
	elif f=='зеленый пр 62':
		f='зеленый пр-кт'
	elif f=='космонавтов 1':
		f='космонавтов у'
	elif f=='локомтивный п':
		f='локомотивный '
	elif f=='лыткарино пар':
		f='лыткарино г,'
	elif f=='матвеевская2.':
		f='матвеевская у'
	elif f=='мичуринский31':
		f='мичуринский п'
	elif f=='ореховый буль':
		f='ореховый б-р,'
	elif f=='отрадная 16.j' or f=='отрадная  16.':
		f='отрадная ул.1'
	elif f=='поречная 10 (':
		f='поречная ул, '
	elif f=='профсоюзная 109 (1)':
		f='профсоюзная ул, д. 109 ('
	photos2+=[re.sub('.jpg','',f)]
	
print(photos2)

wb = load_workbook('./Zott.xlsx')            #бахаем список из таблицы
shoplist=[]
sheet=wb.get_sheet_by_name('Виктория')
for i in range(2,28):
    i=[sheet.cell(row=i, column=4).value]
    shoplist+=i
shoplist2=[]
for j in shoplist:
	j2=j.rstrip().lower()
	if j2[:11]=='профсоюзная':
		j2=j2[:24]
	else:
		j2=j2[:13]
	shoplist2+=[j2]
print(shoplist2)
#for e in shoplist2:
#	print(e)
for g in shoplist2:
	if g not in photos2:
		print('Не хватает фотографий по', g)